
from flask import render_template, redirect, request, session, url_for, send_file
from erp import app, db, bcrypt
from erp.models import User
from flask_login import login_user, current_user, logout_user, login_required
from erp.marks import marksdb, attendancedb
import pandas as pd
import pymysql
from openpyxl import load_workbook
from datetime import date, timedelta


path = r"erp/data.xlsx"


markdata = marksdb('marksfile')
# markdata.add_id('123456')
# markdata.values['123456']['sem1']['sub'] = ['123', '456', '789', '123', '456']
# markdata.values['123456']['sem1']['iat1']['marks'] = [80,60,90,80,80]
# markdata.commit()
attendancedata = attendancedb('attendancefile')
# attendancedata.add_id('123456')
# attendancedata.values['123456']['sem1']['total'] = 50
# attendancedata.values['123456']['sem1']['present'] = 45
# attendancedata.values['123456']['sem1']['absent'] = 5
# attendancedata.values['123456']['sem1']['od'] = 0
attendancedata.commit()


@app.route('/')
@app.route('/home')
def home():
    if current_user.is_authenticated:
        return redirect(url_for('main'))
    else:
        return render_template('home.html', title='Home')


@app.route('/register', methods=['GET', 'POST'])
def register():
    if current_user.is_authenticated:
        return redirect(url_for('main'))
    else:
        hashed_password = bcrypt.generate_password_hash(
            "admin").decode('utf-8')
        user = User(password=hashed_password, uid='admin', usertype='admin')
        db.session.add(user)
        db.session.commit()
        print(user)

    # if validated:
    # return redirect(url_for('home'))


@app.route('/logout')
def logout():
    logout_user()
    return redirect(url_for('home'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('home'))

    elif request.method == 'POST':
        uid = request.form.get('uid')
        password = request.form.get('password')
        user = User.query.filter_by(uid=uid).first()
        print(uid, password)
        print(user.uid)
        if user and bcrypt.check_password_hash(user.password, password):
            print("going in")
            login_user(user)
            print(current_user.uid)
        
        else:
            return render_template('login.html', error="Invalid Credentials")
        return redirect(url_for('home'))
    else:
        return render_template('login.html',)


@app.route('/main')
@login_required
def main():
    session['message'] = ''
    if current_user.usertype == 'student':
        db = pymysql.connect('localhost', password='root',
                             db='erp', user='anupamkris')
        cursor = db.cursor()
        cursor.execute(f'select * from profile where id={current_user.uid}')
        data = cursor.fetchall()[0]
        db.close()
        return render_template('main.html', data=data)
    elif current_user.usertype == 'admin':
        return redirect(url_for('admin'))
    elif current_user.usertype == 'account':
        return redirect(url_for('accountspage'))
    elif current_user.usertype == 'staff':
        return redirect(url_for('staff_home'))
    else:
        return redirect('/attendance/mark')


@app.route('/attendance')
@login_required
def attendance():
    if request.args:
        db = pymysql.connect('localhost', password='root',
                             db='erp', user='anupamkris')
        cursor = db.cursor()
        sem = request.args.get('sem')
        year = request.args.get('year')
        sem = 2 * (int(year) - 1) + int(sem)
        print('\n\n\n\n\n\n\n', sem, year)
        data = attendancedata.values[current_user.uid]['sem'+str(sem)]
        print('\n\n\n\n\n\n\n', data)
        cursor.execute(
            f"select year,branch from profile where id = {current_user.uid}")
        data = cursor.fetchall()[0]
        year = data[0]
        branch = data[1]

        print('\n\\n\n\n\n\n\mn\n\n\n\n', year)
        cursor.execute(
            f'select s{sem}s,s{sem}e from semdates where year = {year}')
        dates = cursor.fetchall()[0]
        print(dates)
        fromdate = str(dates[0]).split('-')
        todate = str(dates[1]).split('-')

        print('\n\n\n\n\n\n\n\n\nFROM DAtE TODATE')
        print(fromdate, todate, year, branch)

        sdate = eval(
            f"date({int(fromdate[0])},{int(fromdate[1])},{int(fromdate[2])})")
        edate = eval(
            f"date({int(todate[0])},{int(todate[1])},{int(todate[2])})")
        delta = edate - sdate
        dates = []
        for i in range(delta.days + 1):
            day = sdate + timedelta(days=i)
            dates.append('-'.join(str(day).split('-')[::-1]))
        print(dates)
        attendance = {}
        try:
            for i in dates:
                df = pd.read_excel(path, sheet_name=i)
                for j in df.values:
                    if j[0] in attendance:
                        attendance[j[0]][j[1]] += 1
                        attendance[j[0]]['t'] += 1
                    else:
                        attendance[j[0]] = {'a': 0, 'p': 0, 'o': 0, 't': 1}
                        attendance[j[0]][j[1]] += 1
                print(attendance)
        except:
            pass
        cursor.execute(f'select id,regno,name from profile')
        stdata = {}
        temp = cursor.fetchall()
        for i in temp:
            stdata[i[0]] = {'regno': i[1], 'name': i[2]}
        print(attendance)
        print('Attemdamce  :  ', attendance[123456])
        return render_template('attendance_view.html', data=attendance[int(current_user.uid)])
    else:
        return render_template('attendance.html')


@app.route('/marks', methods=['GET', 'POST'])
@login_required
def marks():
    if request.args:
        db = pymysql.connect('localhost', password='root',
                             db='erp', user='anupamkris')
        cursor = db.cursor()
        global markdata
        sem = request.args.get('semester')
        test = request.args.get('test')
        print('vals', sem, test, markdata.values[current_user.uid])
        cursor.execute(f'select code, name from subjects')
        data = cursor.fetchall()
        subcode = {}
        for i in data:
            subcode[i[0]] = i[1]
        
        return render_template('marks.html',
                               sublist=markdata.values[current_user.uid][sem]['sub'],
                               markslist=markdata.values[current_user.uid][sem][test]['marks'],
                               zip=zip,
                               subcode=subcode,
                               testname=test.upper())
    else:
        return render_template('markpick.html')


@app.route('/subjects')
@login_required
def subjects():
    if not request.args:
        db = pymysql.connect('localhost', password='root',
                             db='erp', user='anupamkris')
        cursor = db.cursor()
        cursor.execute(
            f'select sem, name from subjects where branch like (select branch from profile where id = {current_user.uid})')
        data = cursor.fetchall()
        # ((1,"sci"),(2,"poi"))
        subject = {}
        for i in data:
            if i[0] in subject:
                subject[i[0]].append(i[1])
            else:
                subject[i[0]] = [0, i[1]]
        print(subject)
        print("\n\n\n\n\\n\n\n\n\n\n\n\\n\nn\n", data)
        cursor.execute(
            f'select * from materials where branch like (select branch from profile where id = {current_user.uid})')
        print('\n\n\n\n\n\\n\n\n\n', subject)
        materials = cursor.fetchall()
        db.close()
        return render_template('subjects.html', subject=subject, materials=materials)
    else:
        subject = request.args.get('subjects')
        type = request.args.get('material')
        sem = request.args.get('sem')
        db = pymysql.connect('localhost', password='root',
                             db='erp', user='anupamkris')
        cursor = db.cursor()
        cursor.execute(
            f'select name, captions, filename from materials where branch like (select branch from profile where id = {current_user.uid}) and subject like (select code from subjects where name like "{subject}") and type like "{type}"')
        data = cursor.fetchall()
        return render_template("materials.html", data=data, type=type)


@app.route('/circular')
@login_required
def circular():
    db = pymysql.connect('localhost', password='root',
                         db='erp', user='anupamkris')
    cursor = db.cursor()
    cursor.execute('select * from circular')
    data = cursor.fetchall()
    db.close()
    return render_template('circular.html', data=data)


@app.route('/fees')
@login_required
def fees():
    db = pymysql.connect('localhost', password='root',
                         db='erp', user='anupamkris')
    cursor = db.cursor()
    cursor.execute(f'select * from fees where id = {current_user.uid}')
    data = cursor.fetchall()[0]
    db.close()
    return render_template('fees.html', data=data)


@app.route('/sts/<filename>')
@login_required
def send_download_file(filename):
    return send_file('sts/materials/' + filename,
                     as_attachment=True,
                     )


@app.route('/sts/notif/<filename>')
@login_required
def send_notif_file(filename):
    return send_file('sts/notifs/' + filename,
                     as_attachment=True,
                     )


@app.route('/profile')
@login_required
def profile():
    db = pymysql.connect('localhost', password='root',
                         db='erp', user='anupamkris')
    cursor = db.cursor()
    cursor.execute(f'select * from profile where id={current_user.uid}')
    data = cursor.fetchall()[0]
    db.close()
    return render_template('profile.html', data=data)

############## A C C O U N T  S ##############


@app.route('/accounts/main')
@login_required
def accountspage():
    if current_user.usertype == 'account':
        return render_template('account/main.html')


@app.route('/accounts/sl')
@login_required
def studentlist():
    if current_user.usertype == 'account':
        # db = pymysql.connect('localhost', password='root',
        #                     db='erp', user='anupamkris')
        # cursor = db.cursor()
        # cursor.execute(
        #     f'select profile.id, name, year, branch, total, paid, balance from profile, fees where profile.id = fees.id')
        # data = cursor.fetchall()
        return render_template('account/student_list.html', 
        # data=data, 
        # len=len
        )


@app.route('/accounts/dl')
@login_required
def duelist():
    if current_user.usertype == 'account':
        # db = pymysql.connect('localhost', password='root',
        #                     db='erp', user='anupamkris')
        # cursor = db.cursor()
        # cursor.execute(
        #     f'select profile.id, name, year, branch, balance from profile, fees where profile.id = fees.id and balance > 0')
        # data = cursor.fetchall()
        return render_template('account/due_list.html', 
        # data=data, 
        # len=len
        )


@app.route('/accounts/scl')
@login_required
def scholarshiplist():
    if current_user.usertype == 'account':
        return render_template('account/scholarship.html')


@app.route('/accounts/ne')
@login_required
def newentry():
    if current_user.usertype == 'account':
        return render_template('account/new_entry.html')

# ------------------------ ADMIN -------------------------------#
#


@app.route('/adminsaddentry', methods=['GET', 'POST'])
@login_required
def adminsaddentry():
    if current_user.usertype == 'admin':
        if request.method == 'POST':
            uid = request.form['uid']
            name = request.form['name']
            year = request.form['year']
            contact = request.form['contact']
            address = request.form['address']
            cutoff = request.form['cutoff']
            brdpoint = request.form['brdpoint']
            area = request.form['area']
            branch = request.form['branch']
            fathername = request.form['fathername']
            fatheroccupation = request.form['fatheroccupation']
            fathercontact = request.form['fathercontact']
            email = request.form['email']
            regno = request.form['regno']
            dob = '-'.join(request.form['dob'].split('-')[::-1])
            db = pymysql.connect('localhost', password='root',
                                 db='erp', user='anupamkris')
            cursor = db.cursor()
            cursor.execute(
                f'insert into profile values ({uid},"{name}",{regno},"{dob}", "{address}", {contact}, "{email}","{brdpoint}", "{branch}", "{area}", "{fathername}", "{fatheroccupation}", {fathercontact}, {cutoff}, {year})')

            db.commit()
            db.close()
            return redirect('/admin')
        else:
            return render_template('admin/saddentry.html')
    else:
        return redirect(url_for('main'))


@app.route('/delete/<studentid>')
@login_required
def deletestudent(studentid):
    if current_user.usertype == 'admin':
        db = pymysql.connect('localhost', password='root',
                             db='erp', user='anupamkris')
        cursor = db.cursor()
        cursor.execute(f'delete from profile where id={studentid}')
        db.commit()
        db.close()
        return redirect(url_for('adminsprofilemain'))
    else:
        pass


@app.route('/admineditprofile', methods=['GET', 'POST'])
@login_required
def admineditprofile():
    if current_user.usertype == 'admin':
        if request.method == 'POST':
            uid = request.form['uid']
            name = request.form['name']
            year = request.form['year']
            contact = request.form['contact']
            address = request.form['address']
            cutoff = request.form['cutoff']
            brdpoint = request.form['brdpoint']
            bloodgrp = request.form['bloodgrp']
            area = request.form['area']
            branch = request.form['branch']
            fathername = request.form['fathername']
            fatheroccupation = request.form['fatheroccupation']
            fathercontact = request.form['fathercontact']
            email = request.form['email']
            regno = request.form['regno']
            dob = '-'.join(request.form['dob'].split('-')[::-1])
            db = pymysql.connect('localhost', password='root',
                                 db='erp', user='anupamkris')
            cursor = db.cursor()
            cursor.execute('select * from profile')
            print(f'update profile set name = "{name}", regno = {regno}, dob = "{dob}", address = "{address}", contact = {contact},email = "{email}",brdpoint = "{brdpoint}", branch = "{branch}",area = "{area}",fathername = "{fathername}",fatheroccupation="{fatheroccupation}",fathercontact={fathercontact}, cutoff={cutoff}, year={year} where id = {uid}')
            cursor.execute(f'update profile set name = "{name}", regno = {regno}, dob = "{dob}", bloodgrp = "{bloodgrp}", address = "{address}", contact = {contact},email = "{email}",brdpoint = "{brdpoint}", branch = "{branch}",area = "{area}",fathername = "{fathername}",fatheroccupation="{fatheroccupation}",fathercontact={fathercontact}, cutoff={cutoff}, year={year} where id = {uid}')
            db.commit()
            db.close()
            return redirect(url_for('adminsprofilemain'))


@app.route('/adminsprofilemain', methods=['GET', 'POST'])
@login_required
def adminsprofilemain():
    if current_user.usertype == 'admin':
        if request.method == 'POST':
            try:
                print('Entering Try Block')
                deleteusername = request.form['deleteuser']
                print(deleteusername)
                db = pymysql.connect(
                    'localhost', password='root', db='erp', user='anupamkris')
                cursor = db.cursor()
                cursor.execute(
                    f'select * from profile where id = {deleteusername}')
                values = cursor.fetchall()[0]
                print(values)
                fieldnames = [i[0] for i in cursor.description]
                data = dict(zip(fieldnames, values))
                return render_template('admin/sdeleteconfirmation.html', data=data, str=str, disabled='disabled', method='delete')
            except:
                try:
                    print('Entering Try Block')
                    deleteusername = request.form['editstudent']
                    print(deleteusername)
                    db = pymysql.connect(
                        'localhost', password='root', db='erp', user='anupamkris')
                    cursor = db.cursor()
                    cursor.execute(
                        f'select * from profile where id = {deleteusername}')
                    values = cursor.fetchall()[0]
                    print(values)
                    fieldnames = [i[0] for i in cursor.description]
                    data = dict(zip(fieldnames, values))
                    return render_template('admin/sdeleteconfirmation.html', data=data, str=str, disabled='', method='edit')
                except:
                    pass
        return render_template('admin/studentprofileentry.html')
    else:

        return redirect(url_for('main'))


@app.route('/adminviewuser')
@login_required
def adminviewuser():
    if current_user.usertype == 'admin':
        users = User.query.all()
        userlist = []
        for i in users:
            if i.usertype == 'student':
                pass
            else:
                userlist.append([i.uid, i.usertype])

        return render_template('admin/viewuser.html', userlist=userlist)
    else:
        return redirect(url_for('main'))


@app.route('/adminviewprofile')
@login_required
def adminviewprofile():
    if current_user.usertype == 'admin':
        db = pymysql.connect('localhost', password='root',
                             db='erp', user='anupamkris')
        cursor = db.cursor()
        cursor.execute(f'select * from profile')
        db.close()
        data = cursor.fetchall()
        return render_template('admin/viewprofile.html', data=data)
    else:
        return redirect(url_for('main'))


@app.route('/admin', methods=['GET', 'POST'])
@login_required
def admin():
    if request.method == 'POST':
        try:
            usertype = request.form['slt-1'].lower()
        except:
            usertype = None
        if usertype:
            username = request.form['username']
            password = request.form['password']
            hashed_password = bcrypt.generate_password_hash(
                password).decode('utf-8')
            user = User(password=hashed_password,
                        uid=username, usertype=usertype)
            db.session.add(user)
            db.session.commit()
            return redirect(url_for('admin'))
        else:
            username = request.form['username']
            user = User.query.filter_by(uid=username).delete()
            db.session.commit()
            print('User Deleted')
    if current_user.usertype == 'admin':
        users = User.query.all()
        userlist = []
        for i in users:
            if i.usertype == 'student':
                pass
            else:
                userlist.append([i.uid, i.usertype])
        return render_template('admin/user.html', userlist=userlist)
    else:
        return redirect(url_for('main'))

# ---------------------- ATTENDANCE ----------------------- #


@app.route('/subject/add', methods=['GET', 'POST'])
@login_required
def subjectadd():
    if current_user.usertype == 'admin':
        if request.method == 'POST':
            sem = request.form['sem']
            subcode = request.form['subcode']
            subname = request.form['subname']
            branch = request.form['branch']
            staffid = request.form['staffid']
            year = request.form['year']
            section = request.form['section']
            db = pymysql.connect('localhost', password='root',
                                 db='erp', user='anupamkris')
            cursor = db.cursor()
            cursor.execute(
                f'insert into subjects values({sem},"{subcode}","{subname}","{branch}","{staffid}",{year},"{section}")')
            db.commit()
            db.close()
            return render_template('attendance/add.html', alert="Subject added successfully")
        else:
            if session['message']:
                alert = session['message']
                session['message'] = ''
            else:
                alert = ''
            return render_template('attendance/add.html', alert=alert)


@app.route('/subject/edit', methods=['GET', 'POST'])
@login_required
def subjectedit():
    if current_user.usertype == 'admin':
        if request.method == 'POST':
            try:
                subname = request.form['subname']
            except:
                subname = None
            if subname:
                xyz = request.form
                sem = request.form['sem']
                subcode = request.form['subcode']
                branch = request.form['branch']
                staffid = request.form['staffid']
                year = request.form['year']
                section = request.form['section']
                db = pymysql.connect('localhost', password='root',
                                     db='erp', user='anupamkris')
                cursor = db.cursor()
                cursor.execute(
                    f'update subjects set sem = {sem},year = "{year}", section = "{section}", branch = "{branch}", staffid = "{staffid}", name = "{subname}" where code = "{subcode}"')
                db.commit()
                db.close()
                session['message'] = 'Edited Successfully'
                return redirect('/subject/add')
            else:
                subcode = request.form['subcode']
                print(subcode)
                db = pymysql.connect('localhost', password='root',
                                     db='erp', user='anupamkris')
                cursor = db.cursor()
                cursor.execute(
                    f'select * from subjects where code = "{subcode}"')
                data = cursor.fetchall()[0]
                return render_template('attendance/edit.html', data=data)
        else:
            return render_template()


@app.route('/attendance/mark', methods=['GET', 'POST'])
@login_required
def markattendance():
    if current_user.usertype == 'attendance':
        if request.method == 'POST':
            session['date'] = request.form['date']
            session['year'] = request.form['year']
            session['branch'] = request.form['branch']
            return redirect(url_for('attendancetable'))
        else:
            if session['message']:
                alert = session['message']
                session['message'] = ''
            else:
                alert = ''
            return render_template('attendance/markattendance.html', alert=alert)


@app.route('/attendance/table', methods=['GET', 'POST'])
@login_required
def attendancetable():
    if current_user.usertype == 'attendance':
        if request.method == 'POST':
            book = load_workbook(path)
            writer = pd.ExcelWriter(path, engine='openpyxl')
            writer.book = book
            try:
                df = pd.read_excel(path, sheet_name=session['date'])
            except:
                df = pd.read_excel(path, sheet_name='template')
            attendance = {}
            db = pymysql.connect('localhost', password='root',
                                 db='erp', user='anupamkris')
            cursor = db.cursor()
            cursor.execute(
                f'select id from profile where branch = \'{session["branch"]}\' and year = {session["year"]}')
            for i in cursor.fetchall():
                df = df.append(
                    {'sid': i[0], 'session1': 'p', 'session2': 'p', 'session3': 'p', 'session4': 'p', 'session5': 'p', 'overall':'p'}, ignore_index=True, )
            print(request.form)
            for i in request.form:
                temp = i.split('-')
                attendance[int(temp[1])] = temp[0][0]
            print(attendance)
            for i in attendance.keys():
                df.loc[df['sid'] == i, ['overall']] = attendance[i]
            print(df)
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            df.to_excel(writer, sheet_name=session['date'], index=False)
            writer.save()
            writer.close()
            session['message'] = "Attendance Marked Sucessfully"
            return redirect(url_for('markattendance'))
        else:
            db = pymysql.connect('localhost', password='root',
                                 db='erp', user='anupamkris')
            cursor = db.cursor()
            cursor.execute(
                f'select id,name from profile where branch = \'{session["branch"]}\' and year = \'{session["year"]}\'')
            data = cursor.fetchall()
            return render_template('attendance/table.html', data=data, len=len)

# insert into semdates values(2018,"2020-10-08","2020-10-12","2020-10-08","2020-10-12","2020-10-08","2020-10-12","2020-10-08","2020-10-12","2020-10-08","2020-10-12","2020-10-08","2020-10-12","2020-10-08","2020-10-12","2020-10-08","2020-10-12")


@app.route('/attendance/report', methods=['GET', 'POST'])
@login_required
def viewreport():
    if current_user.usertype == 'attendance':
        if request.method == 'POST':
            fromdate = (request.form['fromdate'].split('/'))[::-1]
            todate = (request.form['todate'].split('/'))[::-1]
            year = request.form['year']
            branch = request.form['branch']
            section = request.form['section']
            print('\n\n\n\n\n\n\n\n\n')
            print(fromdate, todate, year, branch)

            sdate = eval(
                f"date({int(fromdate[0])},{int(fromdate[1])},{int(fromdate[2])})")
            edate = eval(
                f"date({int(todate[0])},{int(todate[1])},{int(todate[2])})")
            delta = edate - sdate
            dates = []
            for i in range(delta.days + 1):
                day = sdate + timedelta(days=i)
                dates.append('-'.join(str(day).split('-')[::-1]))
            print(dates)
            attendance = {}
            for i in dates:
                try:
                    df = pd.read_excel(path, sheet_name=i)
                    for j in df.values:
                        if j[0] in attendance:
                            attendance[j[0]][j[1]] += 1
                            attendance[j[0]]['t'] += 1
                        else:
                            attendance[j[0]] = {'a': 0, 'p': 0, 'o': 0, 't': 1}
                            attendance[j[0]][j[1]] += 1
                except:
                    pass
                print(attendance)
            db = pymysql.connect('localhost', password='root',
                                 db='erp', user='anupamkris')
            cursor = db.cursor()
            cursor.execute(f'select id,regno,name from profile where section like "{section}"')
            stdata = {}
            temp = cursor.fetchall()
            for i in temp:
                stdata[i[0]] = {'regno': i[1], 'name': i[2]}
            if len(stdata) == 0:
                session['message'] = 'No Data Found'
                return redirect(url_for('viewreport'))
            return render_template('attendance/reportgenerated.html', attendance=attendance, stdata=stdata, len=len, zip=zip, dates=dates)
        else:
            if message:=session['message']:
                session['message'] = ''                
            return render_template('attendance/viewreport.html', message=message)


@app.route('/attendance/viewreport', methods=['GET', 'POST'])
@login_required
def report():
    return render_template('attendance/reportgenerated.html')


#------------------------------------- STAFF SIDE -----------------------------------------#

@app.route('/staff/home', methods= ['GET','POST'])
@login_required
def staff_home():
    if current_user.usertype == 'staff':
        db = pymysql.connect('localhost', password='root',
                             db='erp', user='anupamkris')
        cursor = db.cursor()
        cursor.execute(f'select name, branch, year, sem, code from subjects where staffid like "{current_user.uid}"')
        data = cursor.fetchall()
        print(data)
        db.close()
        return render_template('staff/staff_h_page.html', i = data)
    else:
        return redirect(url_for('main'))

@app.route('/staff/materials', methods= ['GET','POST'])
@login_required
def staff_materials():
    if current_user.usertype == 'staff':
        if request.method == 'POST':
            upfile = request.files['filename'] 
            branch = request.form['branch']
            sem = request.form['sem']
            subject  = request.form['subject']
            mtype = request.form['mtype']
            captions = request.form['captions']
            import os
            upfile.save("./erp/sts/materials/"+upfile.filename)
            db = pymysql.connect('localhost', password='root',
                                db='erp', user='anupamkris')
            cursor = db.cursor()
            cursor.execute(f'select code from subjects where name like "{subject}"')
            subject = cursor.fetchall()[0][0]
            cursor.execute(f'insert into materials values ( "{branch}", "{sem}", "{subject}", "{upfile.filename}", "xyz", "{mtype}", "{captions}")')
            db.commit()
            db.close()
            session['message'] = 'Uploaded Successfully'
            return redirect(url_for('staff_materials'))
        else:
            try:
                print(session['message'])
                if session['message']:
                    alertmessage = session['message']
                    session['message'] = ''
                else:
                    alertmessage = ''
            except:
                alertmessage = ''
            print(alertmessage)
            db = pymysql.connect('localhost', password='root',
                                db='erp', user='anupamkris')
            cursor = db.cursor()
            cursor.execute(f'select * from subjects where staffid like "{current_user.uid}"')
            data = cursor.fetchall()
            cursor.execute(f'select branch from subjects where staffid like "{current_user.uid}" group by branch')
            secdata = cursor.fetchall()
            print(data)
            db.close()
            return render_template('staff/materials.html', data=data, secdata = secdata, alertmessage=alertmessage)

@app.route('/staff_marks', methods=['GET', 'POST'])
@login_required
def staff_marks():
    if current_user.usertype == 'staff':
        if request.method == 'POST':
            session['testname'] = request.form['testname']
            session['sec'] = request.form['sec']
            # session['branch'] = request.form['branch']
            # session['subject'] = request.form['subject']
            session['subject'] = 'CSE1285'
            session['branch'] = 'CSE'
            return redirect(url_for('staff_marks2'))

        else:
            db = pymysql.connect('localhost', password='root',
                                db='erp', user='anupamkris')
            cursor = db.cursor()
            cursor.execute(f'select distinct branch from subjects where staffid like "{current_user.uid}"')
            data = cursor.fetchall()
            cursor.execute(f'select code from subjects where staffid like "{current_user.uid}" group by branch')
            secdata = cursor.fetchall()

            return render_template('staff/marks_1.html', branches=data, subjects=secdata)


@app.route('/staff_marks2', methods=['GET', 'POST'])
@login_required
def staff_marks2():
    if current_user.usertype == 'staff':
        if request.method == 'POST':
            
            data = dict(request.form)
            print(markdata.values)
            for i in data.keys():
                markdata.values[str(i)]['sem'+str(session['sem'])][''.join(session['testname'].lower().split())]['marks'].append(data[i])
                markdata.values[str(i)]['sem'+str(session['sem'])]['sub'].append(session['subject'])
                markdata.commit()
            session.pop("sem", None)
            session.pop("testname", None)
            session.pop("branch", None)
            session.pop("subject", None)
            session.pop("sec", None)
            return redirect(url_for('main'))
        else:
            sec = session['sec']
            testname = session['testname']
            branch = session['branch']
            subject = session['subject']
            db = pymysql.connect('localhost', password='root',
                                db='erp', user='anupamkris')
            cursor = db.cursor()
            cursor.execute(f'select id, name from profile where section like "{sec}" and branch like "{branch}"')
            data = cursor.fetchall()
            cursor.execute(f'select sem from subjects where code like "{subject}"')
            session['sem'] = cursor.fetchall()[0][0]
            db.close()
            return render_template('staff/marks_2.html', data=data)

@app.route('/staff_attendance', methods=['GET', 'POST'])
@login_required
def staff_attendance():
    if current_user.usertype == 'staff':
        if request.method == 'POST':
            session['branch'] = request.form['branch']
            session['year'] = request.form['year']
            session['date'] = '-'.join(request.form['date'].split('-')[::-1])
            session['session'] = request.form['session']
            session['sec'] = request.form['sec']
            
            return redirect(url_for('staff_attendance_mark'))
        else:
            db = pymysql.connect('localhost', password='root',
                                 db='erp', user='anupamkris')
            cursor = db.cursor()
            cursor.execute(f'select branch from subjects where staffid like "{current_user.uid}"')
            data = cursor.fetchall()[0]
            return render_template('staff/attendance.html', data =data)


@app.route('/staff_attendance/mark', methods=['GET', 'POST'])
@login_required
def staff_attendance_mark():
    if current_user.usertype == 'staff':
        if request.method == 'POST':
            book = load_workbook(path)
            writer = pd.ExcelWriter(path, engine='openpyxl')
            writer.book = book
            try:
                df = pd.read_excel(path, sheet_name=session['date'])
            except:
                df = pd.read_excel(path, sheet_name='template')
            attendance = {}
            db = pymysql.connect('localhost', password='root',
                                 db='erp', user='anupamkris')
            cursor = db.cursor()
            cursor.execute(
                f'select id from profile where branch = \'{session["branch"]}\' and year = {session["year"]} and section like \'{session["sec"]}\'')
            for i in cursor.fetchall():
                for j in df.values:
                    if i[0] in j:
                        tempvar = True
                        break
                else:
                    tempvar=False
                if tempvar:
                    pass
                else:
                    df = df.append(
                        {'sid': i[0], 'session1': 'p', 'session2': 'p', 'session3': 'p', 'session4': 'p', 'session5': 'p', 'overall':'p'}, ignore_index=True, )
            cdata = request.form
            for i in cdata:
                temp = i.split('-')
                attendance[int(temp[1])] = temp[0][0]
            print(attendance)
            for i in attendance.keys():
                df.loc[df['sid'] == i, [session['session']]] = attendance[i]

            print(df)
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            df.to_excel(writer, sheet_name=session['date'], index=False)
            writer.save()
            writer.close()
            session['message'] = "Attendance Marked Sucessfully"
            db.close()
            return redirect(url_for('main'))
        else:
            db = pymysql.connect('localhost', password='root',
                                 db='erp', user='anupamkris')
            cursor = db.cursor()
            cursor.execute(
                f'select id,name from profile where branch = \'{session["branch"]}\' and year = \'{session["year"]}\' and section = \'{session["sec"]}\'')
            data = cursor.fetchall()
            print(data)
            db.close()
            return render_template('staff/markAttendance.html', data=data, len=len)
